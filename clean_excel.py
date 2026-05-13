"""
Villcan - Clean Excel data and generate CSV files for Supabase UI import
This script:
1. Reads the Excel file (BD- Villcan.xlsx)
2. Cleans and normalizes all values (fixes ×10^13 scaling)
3. Generates 3 CSV files: contacts.csv, services.csv, movements.csv
"""

import openpyxl
import csv
import re
from datetime import datetime

EXCEL_FILE = 'C:/Users/User/Documents/proyectos/villcan/BD- Villcan.xlsx'
CONTACTS_CSV = 'C:/Users/User/Documents/proyectos/villcan/villcan/contacts.csv'
SERVICES_CSV = 'C:/Users/User/Documents/proyectos/villcan/villcan/services.csv'
MOVEMENTS_CSV = 'C:/Users/User/Documents/proyectos/villcan/villcan/movements.csv'

# Service prices (from servicios sheet + inferred)
SERVICE_PRICES = {
    'corte clasico': 40000,
    'corte clásico': 40000,
    'degradado': 45000,
    'corte niño': 35000,
    'corte niño': 35000,
    'barba': 30000,
    'combo c,b,c': 70000,
    'corte y barba': 60000,
    'corte clásico y barba': 60000,
    'corte y ceja': 55000,
}


def parse_price_from_monto(monto_str):
    """Parse 'USD\xa040000.00' -> 40000"""
    if not monto_str:
        return None
    nums = re.findall(r'[\d,]+', str(monto_str).replace('USD', '').replace('\xa0', ''))
    if nums:
        return int(nums[0].replace(',', ''))
    return None


def get_service_price(service_name):
    """Get price for service, or infer from name"""
    if not service_name:
        return None
    name = service_name.lower().strip()
    if name in SERVICE_PRICES:
        return SERVICE_PRICES[name]
    # Try partial match
    for key in SERVICE_PRICES:
        if key in name or name in key:
            return SERVICE_PRICES[key]
    return None


def clean_amount(value):
    """Clean amount from Excel - removes ×10^13 artifact"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        if value < 0:
            # Negative scaled value: divide by 10^13 and take absolute
            return int(abs(value / 10**13))
        if abs(value) >= 10_000_000_000:  # >= 10^10 means likely scaled
            return int(value / 10**13)
        return int(value)
    if isinstance(value, str):
        value = value.replace('USD', '').replace('\xa0', '').replace(' ', '')
        try:
            return int(float(value.replace(',', '')))
        except:
            return 0
    return 0


def clean_expense(value):
    """Calculate expense from scaled Excel value"""
    if not value or value == 0:
        return 0
    if isinstance(value, (int, float)):
        if value < 0:
            # Negative scaled value: divide by 10^13 and take absolute
            return int(abs(value / 10**13))
        if abs(value) >= 10_000_000_000:  # >= 10^10 means likely scaled
            return int(value / 10**13)
    return int(value) if isinstance(value, (int, float)) else 0


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Movimientos de Caja']

    # Collect unique contacts
    all_contacts = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name and isinstance(name, str):
            all_contacts.add(name.strip())

    # Collect unique services from movements
    all_services = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        svc = row[1]
        if svc and isinstance(svc, str):
            all_services.add(svc.strip())

    # Build services lookup: name -> UUID (we'll use name-based IDs)
    service_uuid_map = {}
    for svc_name in sorted(all_services):
        price = get_service_price(svc_name)
        if price:
            # Create a deterministic ID based on service name
            safe_name = svc_name.lower().replace(' ', '_').replace('ñ', 'n').replace(',', '')
            service_uuid_map[svc_name.strip().lower()] = safe_name

    # Build contacts lookup: name -> UUID
    contact_uuid_map = {}
    contact_list = []
    for contact_name in sorted(all_contacts):
        safe_name = contact_name.lower().strip().replace(' ', '_')
        contact_uuid_map[contact_name.strip().lower()] = safe_name
        contact_list.append(contact_name)

    # ========== WRITE SERVICES CSV ==========
    with open(SERVICES_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'name', 'price', 'is_active'])
        for svc_name in sorted(all_services):
            price = get_service_price(svc_name)
            if price:
                safe_id = service_uuid_map[svc_name.strip().lower()]
                writer.writerow([safe_id, svc_name, price, 'true'])

    # ========== WRITE CONTACTS CSV ==========
    with open(CONTACTS_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'full_name'])
        for contact_name in sorted(all_contacts):
            safe_id = contact_uuid_map[contact_name.strip().lower()]
            writer.writerow([safe_id, contact_name])

    # ========== WRITE MOVEMENTS CSV ==========
    with open(MOVEMENTS_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'id', 'type', 'amount_charged', 'income', 'expense',
            'payment_method', 'contact_id', 'service_id', 'comment', 'created_at'
        ])

        for row in ws.iter_rows(min_row=2, values_only=True):
            tipo = row[2]
            if not tipo:
                continue

            tipo = tipo.lower().strip()

            # Common fields
            income_raw = row[4] if row[4] else 0
            expense_raw = row[5] if row[5] else 0
            comment = row[7]
            fecha = row[9]

            # Parse amount using clean_amount (handles ×10^13 scaling)
            income = clean_amount(income_raw)

            # Parse method
            method = row[6]
            if method:
                method = method.lower().strip()
                if 'efectivo' in method:
                    method = 'efectivo'
                elif 'transfer' in method:
                    method = 'transferencia'
                elif 'pos' in method.lower():
                    method = 'pos'
                else:
                    method = ''
            else:
                method = ''

            # Format date
            if fecha:
                if hasattr(fecha, 'strftime'):
                    fecha_str = fecha.strftime('%Y-%m-%d')
                else:
                    fecha_str = str(fecha)[:10]
            else:
                fecha_str = '2025-01-01'

            # Clean comment
            comment_str = (comment or '').replace("'", "''")

            if tipo == 'servicio':
                service_name = row[1] if row[1] else None
                contact_name = row[0] if row[0] else None
                monto_cobrar = parse_price_from_monto(row[3])

                # Use monto_cobrar as income for servicio
                income = monto_cobrar if monto_cobrar else income

                # Clean expense (vuelto)
                expense = clean_expense(expense_raw)

                # Get service_id
                service_id = ''
                if service_name:
                    svc_lower = service_name.strip().lower()
                    service_id = service_uuid_map.get(svc_lower, '')

                # Get contact_id
                contact_id = ''
                if contact_name:
                    contact_key = contact_name.strip().lower()
                    contact_id = contact_uuid_map.get(contact_key, '')

                writer.writerow([
                    '', 'servicio', income, income, expense,
                    method, contact_id, service_id, comment_str, fecha_str
                ])

            elif tipo in ('gasto', 'apertura', 'cierre'):
                writer.writerow([
                    '', tipo, '', income, 0,
                    '', '', '', comment_str, fecha_str
                ])

    # ========== SUMMARY ==========
    print(f'Generated CSV files:')
    print(f'  - {SERVICES_CSV}  ({len(service_uuid_map)} services)')
    print(f'  - {CONTACTS_CSV}  ({len(contact_uuid_map)} contacts)')

    # Count movements
    movement_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            movement_count += 1
    print(f'  - {MOVEMENTS_CSV}  ({movement_count} movements)')


if __name__ == '__main__':
    main()