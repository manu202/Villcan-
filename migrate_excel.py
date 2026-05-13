"""
Villcan - Migrate data from Excel to Supabase

Usage:
    python migrate_excel.py

Requirements:
    pip install openpyxl psycopg2-binary

This script:
1. Reads BD- Villcan.xlsx
2. Cleans and transforms data (fixes the ×10^13 issue)
3. Inserts into Supabase
"""

import openpyxl
from supabase import create_client, Client
import sys

# Supabase connection
SUPABASE_URL = "https://mhlkrvxcgrntiuglufdl.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im1obGtydnhjZ3JudGl1Z2x1ZmRsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg2Mzc1MzAsImV4cCI6MjA5NDIxMzUzMH0.TKeGS68VyBgfkYQGzUD45kVSiHKXRckA387BYYgZ3q0"

# Excel file path (relative to project root)
EXCEL_FILE = "../BD- Villcan.xlsx"

def clean_amount(value):
    """Clean amount from Excel - removes ×10^13 artifact"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        # The numbers in Excel are ×10^13, divide to get real value
        if value > 1_000_000_000_000:  # Likely the artifact
            return int(value / 10**13)
        return int(value)
    if isinstance(value, str):
        # Remove 'USD' and spaces, handle formatting
        value = value.replace('USD', '').replace('\xa0', '').replace(' ', '')
        try:
            return int(float(value.replace(',', '')))
        except:
            return 0
    return 0

def clean_string(value):
    """Clean string values"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return str(value) if value else None

def get_or_create_contact(supabase, name):
    """Get contact by name or create if doesn't exist"""
    # Clean name
    name = clean_string(name)
    if not name:
        return None

    # Try to find existing
    response = supabase.table('contacts').select('id, full_name').ilike('full_name', name).execute()
    if response.data:
        return response.data[0]['id']

    # Create new
    response = supabase.table('contacts').insert({
        'full_name': name
    }).execute()

    if response.data:
        return response.data[0]['id']
    return None

def get_or_create_service(supabase, name):
    """Get service by name or create if doesn't exist"""
    name = clean_string(name)
    if not name:
        return None

    # Try to find existing
    response = supabase.table('services').select('id, name').ilike('name', name).execute()
    if response.data:
        return response.data[0]['id']

    # Create new with default price (user should verify)
    response = supabase.table('services').insert({
        'name': name,
        'price': 35000  # Default price
    }).execute()

    if response.data:
        return response.data[0]['id']
    return None

def migrate_contacts(supabase, wb):
    """Migrate contacts from Clientes sheet"""
    print("\n[MIGRATING] Contacts...")

    ws = wb['Clientes']
    headers = [cell.value for cell in ws[1]]

    contacts_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        full_name = clean_string(row[0])
        if not full_name:
            continue

        contact = {
            'full_name': full_name,
            'ci': clean_string(row[1]) if len(row) > 1 else None,
            'phone': clean_string(row[2]) if len(row) > 2 else None,
            'comment': clean_string(row[3]) if len(row) > 3 else None,
        }
        contacts_data.append(contact)

    if contacts_data:
        # Insert in batches
        batch_size = 50
        for i in range(0, len(contacts_data), batch_size):
            batch = contacts_data[i:i+batch_size]
            response = supabase.table('contacts').insert(batch).execute()
            print(f"  Inserted {len(response.data)} contacts (batch {i//batch_size + 1})")

    print(f"  [OK] Total contacts: {len(contacts_data)}")

def migrate_services(supabase, wb):
    """Migrate services from servicios sheet"""
    print("\n[MIGRATING] Services...")

    ws = wb['servicios']
    headers = [cell.value for cell in ws[1]]

    services_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        name = clean_string(row[0])
        if not name:
            continue

        # Price is in format 'USD 40000.00' or as number
        price = 35000
        if len(row) > 1 and row[1]:
            price = clean_amount(row[1])

        service = {
            'name': name,
            'price': price,
            'is_active': True
        }
        services_data.append(service)

    if services_data:
        response = supabase.table('services').insert(services_data).execute()
        print(f"  [OK] Inserted {len(response.data)} services")

def migrate_movements(supabase, wb):
    """Migrate movements from Movimientos de Caja sheet"""
    print("\n[MIGRATING] Movements...")

    ws = wb['Movimientos de Caja']
    headers = [cell.value for cell in ws[1]]

    # Map column indices
    col_map = {
        'cliente': 0,
        'servicio': 1,
        'tipomovimiento': 2,
        'monto_cobrar': 3,
        'ingreso': 4,
        'egreso': 5,
        'metodo_pago': 6,
        'comentario': 7,
        'fecha': 9,
        'hora': 10,
        'fuente': 11,
    }

    movements_data = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            tipo = clean_string(row[col_map['tipomovimiento']])
            if not tipo:
                continue

            # Get contact and service IDs
            contact_name = clean_string(row[col_map['cliente']]) if row[col_map['cliente']] else None
            service_name = clean_string(row[col_map['servicio']]) if row[col_map['servicio']] else None

            # Parse tipo
            tipo = tipo.lower().strip()
            tipo_map = {
                'servicio': 'servicio',
                'gasto': 'gasto',
                'apertura': 'apertura',
                'cierre': 'cierre'
            }
            tipo = tipo_map.get(tipo, 'servicio')

            # Parse payment method
            method = clean_string(row[col_map['metodo_pago']])
            if method:
                method = method.lower().strip()
                if 'efectivo' in method:
                    method = 'efectivo'
                elif 'transfer' in method:
                    method = 'transferencia'
                elif 'pos' in method.lower():
                    method = 'pos'
                else:
                    method = None
            else:
                method = None

            # Parse amounts
            amount_charged = clean_amount(row[col_map['monto_cobrar']]) if row[col_map['monto_cobrar']] else None
            income = clean_amount(row[col_map['ingreso']]) if row[col_map['ingreso']] else 0
            expense = clean_amount(row[col_map['egreso']]) if row[col_map['egreso']] else 0

            # Parse date
            fecha = row[col_map['fecha']]
            created_at = None
            if fecha:
                if hasattr(fecha, 'isoformat'):
                    created_at = fecha.isoformat()
                else:
                    created_at = str(fecha)

            movement = {
                'type': tipo,
                'amount_charged': amount_charged if amount_charged > 0 else None,
                'income': income,
                'expense': expense,
                'payment_method': method,
                'comment': clean_string(row[col_map['comentario']]) if len(row) > col_map['comentario'] else None,
                'created_at': created_at,
            }

            # For servicio, we need contact_id and service_id
            if tipo == 'servicio':
                if contact_name:
                    # This will be resolved after contacts are migrated
                    movement['_contact_name'] = contact_name
                if service_name:
                    movement['_service_name'] = service_name

            movements_data.append(movement)

        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")

    # Now resolve contact and service IDs
    print("  Resolving contact and service IDs...")

    for movement in movements_data:
        contact_name = movement.pop('_contact_name', None)
        service_name = movement.pop('_service_name', None)

        if contact_name:
            response = supabase.table('contacts').select('id').ilike('full_name', contact_name).execute()
            if response.data:
                movement['contact_id'] = response.data[0]['id']

        if service_name:
            response = supabase.table('services').select('id').ilike('name', service_name).execute()
            if response.data:
                movement['service_id'] = response.data[0]['id']

    # Insert in batches
    if movements_data:
        batch_size = 50
        total_inserted = 0
        for i in range(0, len(movements_data), batch_size):
            batch = movements_data[i:i+batch_size]
            try:
                response = supabase.table('movements').insert(batch).execute()
                total_inserted += len(response.data)
                print(f"  Inserted {len(response.data)} movements (batch {i//batch_size + 1})")
            except Exception as e:
                print(f"  Batch error: {e}")

        print(f"  ✓ Total movements: {len(movements_data)}")

    if errors:
        print(f"  ⚠️ {len(errors)} errors (see below)")
        for err in errors[:5]:
            print(f"    {err}")

def main():
    print("Villcan Migration Tool")
    print("=" * 40)

    # Initialize Supabase client
    print("\nConnecting to Supabase...")
    try:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("  [OK] Connected!")
    except Exception as e:
        print(f"  [FAIL] Connection failed: {e}")
        sys.exit(1)

    # Read Excel
    print(f"\nReading Excel: {EXCEL_FILE}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        print(f"  [OK] Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"  [FAIL] Failed to read Excel: {e}")
        sys.exit(1)

    # Run migrations
    print("\n" + "=" * 40)
    print("Starting migration...")
    print("=" * 40)

    # 1. Contacts first (needed for movements)
    migrate_contacts(supabase, wb)

    # 2. Services
    migrate_services(supabase, wb)

    # 3. Movements
    migrate_movements(supabase, wb)

    print("\n" + "=" * 40)
    print("Migration complete!")
    print("=" * 40)

    # Initialize Supabase client
    print("\nConnecting to Supabase...")
    try:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("  ✓ Connected!")
    except Exception as e:
        print(f"  ✗ Connection failed: {e}")
        sys.exit(1)

    # Read Excel
    print(f"\nReading Excel: {EXCEL_FILE}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        print(f"  ✓ Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"  ✗ Failed to read Excel: {e}")
        sys.exit(1)

    # Run migrations
    print("\n" + "=" * 40)
    print("Starting migration...")
    print("=" * 40)

    # 1. Contacts first (needed for movements)
    migrate_contacts(supabase, wb)

    # 2. Services
    migrate_services(supabase, wb)

    # 3. Movements
    migrate_movements(supabase, wb)

    print("\n" + "=" * 40)
    print("✅ Migration complete!")
    print("=" * 40)

if __name__ == "__main__":
    main()