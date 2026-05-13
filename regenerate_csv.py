"""
Villcan - Regenerate CSV files with ALL columns matching Excel + DB schema

Usage:
    python regenerate_csv.py
"""

import openpyxl
import csv
import re
from datetime import datetime

EXCEL_FILE = 'C:/Users/User/Documents/proyectos/villcan/BD- Villcan.xlsx'
OUTPUT_DIR = 'C:/Users/User/Documents/proyectos/villcan/villcan/'

# Service prices (from servicios sheet + inferred)
SERVICE_PRICES = {
    'corte clasico': 40000,
    'corte clásico': 40000,
    'degradado': 45000,
    'corte niño': 35000,
    'barba': 30000,  # From Excel servicios sheet
    'combo c,b,c': 70000,
    'corte y barba': 60000,
    'corte clásico y barba': 60000,
    'corte y ceja': 55000,
}


def parse_price_from_monto(monto_str):
    """Parse 'USD 35000.00' -> 35000"""
    if not monto_str:
        return None
    # Handle 'USD\xa040000.00' format (non-breaking space)
    nums = re.findall(r'[\d,]+', str(monto_str).replace('USD', '').replace('\xa0', '').replace(' ', ''))
    if nums:
        return int(nums[0].replace(',', ''))
    return None


def get_service_price(service_name):
    """Get price for service, or infer from name"""
    if not service_name:
        return None
    name = service_name.lower().strip()
    if name in SERVICE_PRICES:
        return SERVICE_PRICES[name]
    # Try partial match
    for key in SERVICE_PRICES:
        if key in name or name in key:
            return SERVICE_PRICES[key]
    return None


def clean_amount(value):
    """Clean amount from Excel - removes ×10^13 artifact"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        if value > 1_000_000_000_000:
            return int(value / 10**13)
        return int(value)
    if isinstance(value, str):
        value = value.replace('USD', '').replace('\xa0', '').replace(' ', '')
        try:
            return int(float(value.replace(',', '')))
        except:
            return 0
    return 0


def normalize_payment_method(method):
    """Normalize payment method to: efectivo, transferencia, pos"""
    if not method:
        return ''
    method = method.lower().strip()
    if 'efectivo' in method:
        return 'efectivo'
    elif 'transfer' in method:
        return 'transferencia'
    elif 'pos' in method.lower():
        return 'pos'
    return ''


def normalize_tipo(tipo):
    """Normalize tipo to lowercase: servicio, gasto, apertura, cierre"""
    if not tipo:
        return ''
    tipo = tipo.lower().strip()
    if tipo in ('servicio', 'gasto', 'apertura', 'cierre'):
        return tipo
    return ''


def make_deterministic_id(name):
    """Create deterministic string ID from name"""
    if not name:
        return ''
    return name.lower().strip().replace(' ', '_').replace('ñ', 'n').replace(',', '')


def format_date(value):
    """Format date as YYYY-MM-DD"""
    if not value:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)[:10]


def main():
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ========== LOAD ALL DATA ==========
    print("Loading data from sheets...")

    # Load Clientes sheet for contacts
    clientes_ws = wb['Clientes']
    clientes_data = []
    for row in clientes_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            clientes_data.append({
                'full_name': row[0],
                'ci': row[1] if len(row) > 1 else None,
                'phone': row[2] if len(row) > 2 else None,
                'comment': row[3] if len(row) > 3 else None,
                'created_at': format_date(row[4]) if len(row) > 4 and row[4] else ''
            })
    print(f"  Loaded {len(clientes_data)} clientes")

    # Load servicios sheet for services
    servicios_ws = wb['servicios']
    servicios_data = []
    for row in servicios_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            price = clean_amount(row[1]) if len(row) > 1 and row[1] else None
            servicios_data.append({
                'name': row[0],
                'price': price
            })
    print(f"  Loaded {len(servicios_data)} servicios")

    # Load Movimientos de Caja sheet
    movimientos_ws = wb['Movimientos de Caja']
    movimientos_data = []
    for row in movimientos_ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # Skip rows without tipo
            movimientos_data.append({
                'cliente': row[0],
                'servicio': row[1],
                'tipomovimiento': row[2],
                'monto_cobrar': row[3],
                'ingreso': row[4],
                'egreso': row[5],
                'metodo_pago': row[6],
                'comentario': row[7],
                'fecha': row[9],
                'hora': row[10],
                'fuente': row[11]
            })
    print(f"  Loaded {len(movimientos_data)} movimientos")

    # ========== BUILD LOOKUP MAPS ==========
    print("Building lookup maps...")

    # Services map: name -> id
    service_id_map = {}
    for svc in servicios_data:
        name = svc['name']
        price = svc['price'] or get_service_price(name)
        if price:
            safe_id = make_deterministic_id(name)
            service_id_map[name.lower().strip()] = safe_id
            service_id_map[name.lower().strip().replace('ñ', 'n')] = safe_id

    # Also add from movements sheet services
    for row in movimientos_ws.iter_rows(min_row=2, values_only=True):
        svc_name = row[1]
        if svc_name and svc_name.strip():
            name_lower = svc_name.strip().lower()
            if name_lower not in service_id_map:
                price = get_service_price(svc_name)
                if price:
                    safe_id = make_deterministic_id(svc_name)
                    service_id_map[name_lower] = safe_id
                    # Also handle ñ variations
                    name_lower_n = name_lower.replace('ñ', 'n')
                    existing_id = service_id_map.get(name_lower_n)
                    if existing_id:
                        service_id_map[name_lower] = existing_id
                    else:
                        service_id_map[name_lower_n] = safe_id

    print(f"  Mapped {len(service_id_map)} services")

    # Contacts map: full_name -> id
    contact_id_map = {}
    for cliente in clientes_data:
        name = cliente['full_name']
        if name:
            safe_id = make_deterministic_id(name)
            contact_id_map[name.strip().lower()] = safe_id

    print(f"  Mapped {len(contact_id_map)} contacts")

    # ========== WRITE SERVICES CSV ==========
    print("\nWriting services.csv...")
    services_csv_path = OUTPUT_DIR + 'services.csv'
    with open(services_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'name', 'price', 'is_active'])

        # Write services from servicios sheet
        for svc in servicios_data:
            name = svc['name']
            price = svc['price'] or get_service_price(name)
            if price:
                safe_id = make_deterministic_id(name)
                writer.writerow([safe_id, name, price, 'true'])

    print(f"  Wrote {len(servicios_data)} services")

    # ========== WRITE CONTACTS CSV ==========
    print("\nWriting contacts.csv...")
    contacts_csv_path = OUTPUT_DIR + 'contacts.csv'
    with open(contacts_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'full_name', 'ci', 'phone', 'comment', 'created_at'])

        for cliente in clientes_data:
            name = cliente['full_name']
            if name:
                safe_id = make_deterministic_id(name)
                ci = cliente['ci'] or ''
                phone = cliente['phone'] or ''
                comment = (cliente['comment'] or '').replace('"', '""')
                created = cliente['created_at'] or ''
                writer.writerow([safe_id, name, ci, phone, comment, created])

    print(f"  Wrote {len(clientes_data)} contacts")

    # ========== WRITE MOVEMENTS CSV ==========
    print("\nWriting movements.csv...")
    movements_csv_path = OUTPUT_DIR + 'movements.csv'
    with open(movements_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'id', 'type', 'amount_charged', 'income', 'expense',
            'payment_method', 'contact_id', 'service_id', 'user_id', 'comment', 'created_at'
        ])

        for mov in movimientos_data:
            tipo = normalize_tipo(mov['tipomovimiento'])
            if not tipo:
                continue

            # Parse amounts
            monto_cobrar = parse_price_from_monto(mov['monto_cobrar'])
            ingreso_raw = mov['ingreso'] if mov['ingreso'] else 0
            egreso_raw = mov['egreso'] if mov['egreso'] else 0

            # Clean income
            income = clean_amount(ingreso_raw)
            expense = clean_amount(egreso_raw)

            # Payment method
            payment_method = normalize_payment_method(mov['metodo_pago'])

            # Build comment with hora and fuente
            comment_parts = []
            if mov['comentario']:
                comment_parts.append(str(mov['comentario']))
            if mov['hora']:
                comment_parts.append(f"Hora: {mov['hora']}")
            if mov['fuente']:
                comment_parts.append(f"Fuente: {mov['fuente']}")
            comment = ' | '.join(comment_parts) if comment_parts else ''

            # Date
            created_at = format_date(mov['fecha'])

            # For servicio type, get contact_id and service_id
            contact_id = ''
            service_id = ''
            amount_charged = ''

            if tipo == 'servicio':
                # Get contact
                cliente_name = mov['cliente']
                if cliente_name:
                    contact_key = cliente_name.strip().lower()
                    contact_id = contact_id_map.get(contact_key, '')

                # Get service
                svc_name = mov['servicio']
                if svc_name:
                    svc_key = svc_name.strip().lower()
                    service_id = service_id_map.get(svc_key, '')
                    # Also try with ñ -> n
                    if not service_id:
                        service_id = service_id_map.get(svc_key.replace('ñ', 'n'), '')

                # Use monto_cobrar as amount_charged
                amount_charged = monto_cobrar if monto_cobrar else ''

                # Income should come from amount_charged for servicio type
                if monto_cobrar:
                    income = monto_cobrar

            # user_id is empty string per user instruction
            user_id = ''

            writer.writerow([
                '',  # id - empty, Supabase will generate UUID
                tipo,
                amount_charged,
                income,
                expense,
                payment_method,
                contact_id,
                service_id,
                user_id,
                comment,
                created_at
            ])

    print(f"  Wrote {len(movimientos_data)} movements")

    # ========== SUMMARY ==========
    print("\n" + "=" * 50)
    print("CSV files regenerated successfully!")
    print(f"  - {services_csv_path}")
    print(f"  - {contacts_csv_path}")
    print(f"  - {movements_csv_path}")
    print("=" * 50)


if __name__ == '__main__':
    main()